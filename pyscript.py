import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

def load_csv_file():
    file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    if file_path:
        return pd.read_csv(file_path), file_path
    return None, None

def convert_and_sort(df, column_name):
    # remove dollar signs and commas, then convert to numeric
    df[column_name] = df[column_name].replace('[$,]', '', regex=True).astype(float)
    sorted_df = df.sort_values(by=column_name, ascending=False)
    return sorted_df[column_name].values, sorted_df.index

def create_matchbooks(our_df, bank_df, our_columns, bank_columns, ours_is_credit):
    our_value_column, bank_value_column = ('Credits', 'Debits') if ours_is_credit else ('Debits', 'Credits')
    our_values, our_indices = convert_and_sort(our_df, our_value_column)
    bank_values, bank_indices = convert_and_sort(bank_df, bank_value_column)

    results = []
    our_index = bank_index = 0

    while our_index < len(our_values) and bank_index < len(bank_values):
        our_val = our_values[our_index]
        bank_val = bank_values[bank_index]

        if our_val == bank_val:
            if our_val == 0:
                break
            our_row = our_df.loc[our_indices[our_index]]     # TODO is this robust to identical values ?
            bank_row = bank_df.loc[bank_indices[bank_index]] # ^
            result_row = {
                'Our_Value': our_val,
                'Bank_Value': bank_val,
                'Match': 'MATCH'
            }
            for col in our_columns:
                if col not in ['Credits', 'Debits']:
                    result_row[f'Our_{col}'] = our_row[our_columns[col]]
            for col in bank_columns:
                if col not in ['Credits', 'Debits']:
                    result_row[f'Bank_{col}'] = bank_row[bank_columns[col]]
            results.append(result_row)
            our_index += 1
            bank_index += 1
        elif our_val > bank_val:
            our_row = our_df.loc[our_indices[our_index]]
            result_row = {
                'Our_Value': our_val,
                'Bank_Value': 'XXX',
                'Match': 'MISMATCH'
            }
            for col in our_columns:
                if col not in ['Credits', 'Debits']:
                    result_row[f'Our_{col}'] = our_row[our_columns[col]]
            for col in bank_columns:
                if col not in ['Credits', 'Debits']:
                    result_row[f'Bank_{col}'] = 'XXX'
            results.append(result_row)
            our_index += 1
        else:
            bank_row = bank_df.loc[bank_indices[bank_index]]
            result_row = {
                'Our_Value': 'XXX',
                'Bank_Value': bank_val,
                'Match': 'MISMATCH'
            }
            for col in our_columns:
                if col not in ['Credits', 'Debits']:
                    result_row[f'Our_{col}'] = 'XXX'
            for col in bank_columns:
                if col not in ['Credits', 'Debits']:
                    result_row[f'Bank_{col}'] = bank_row[bank_columns[col]]
            results.append(result_row)
            bank_index += 1

    # Remaining entries from our_value_column
    while our_index < len(our_values):
        if our_values[our_index] == 0: break
        our_row = our_df.loc[our_indices[our_index]]
        result_row = {
            'Our_Value': our_row[our_column],
            'Bank_Value': 'XXX',
            'Match': 'MISMATCH'
        }
        for col in our_columns:
            if col not in ['Credits', 'Debits']:
                result_row[f'Our_{col}'] = our_row[our_columns[col]]
        for col in bank_columns:
            if col not in ['Credits', 'Debits']:
                result_row[f'Bank_{col}'] = 'XXX'
        results.append(result_row)
        our_index += 1

    # Remaining entries from bank_value_column
    while bank_index < len(bank_values):
        if bank_values[bank_index] == 0: break
        bank_row = bank_df.loc[bank_indices[bank_index]]
        result_row = {
            'Our_Value': 'XXX',
            'Bank_Value': bank_row[bank_column],
            'Match': 'MISMATCH'
        }
        for col in our_columns:
            if col not in ['Credits', 'Debits']:
                result_row[f'Our_{col}'] = 'XXX'
        for col in bank_columns:
            if col not in ['Credits', 'Debits']:
                result_row[f'Bank_{col}'] = bank_row[bank_columns[col]]
        results.append(result_row)
        bank_index += 1

    return pd.DataFrame(results)

def save_to_excel_with_color(df, filename):
    wb = Workbook()
    ws = wb.active

    green_fill = PatternFill(start_color='00C851', end_color='00C851', fill_type='solid')
    red_fill = PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid')

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws.append(row)
        if r_idx==0:
            continue
        row_color = green_fill if row[2] == 'MATCH' else red_fill
        for cell in ws[r_idx + 1]:
            cell.fill = row_color


    ws.delete_cols(3)
    wb.save(filename)
    print("Saved " + filename)

import tkinter as tk
from tkinter import filedialog
import pandas as pd
import re
import os

class CSVMatcherApp:
    our_file_name = ""
    bank_file_name = ""
    file_month = ""
    def __init__(self, root):
        self.root = root
        self.root.title("CSV Matcher")

        self.our_csv = None
        self.bank_csv = None

        self.our_file_btn = tk.Button(root, text="Load OUR CSV", command=self.load_our_csv)
        self.our_file_btn.pack()

        self.bank_file_btn = tk.Button(root, text="Load BANK CSV", command=self.load_bank_csv)
        self.bank_file_btn.pack()

        self.match_btn = tk.Button(root, text="Match CSVs", command=self.match_csvs, state=tk.DISABLED)
        self.match_btn.pack()

        self.frame = tk.Frame(root)
        self.frame.pack(pady=20)

        self.our_col_label = tk.Label(self.frame, text="SAGE COLUMNS")
        self.our_col_label.grid(row=0, column=0, padx=5)

        self.bank_col_label = tk.Label(self.frame, text="BANK COLUMNS")
        self.bank_col_label.grid(row=0, column=2, padx=5)

        self.our_entries = []
        self.bank_entries = []

    def load_our_csv(self):
        self.our_csv, temp = load_csv_file()
        self.our_file_name = os.path.basename(temp)
        if self.our_csv is not None:
           month_pattern = r'\b(January|February|March|April|May|June|July|August|September|October|November|December)\b'
           match1 = re.search(month_pattern, self.our_file_name, re.IGNORECASE)
           match2 = re.search(month_pattern, self.bank_file_name, re.IGNORECASE)
           if match1 and match2 and match1.group(0) == match2.group(0):
             self.file_month=match1.group(0)
           self.display_columns(self.our_csv, 'our')
           self.check_files_loaded()

    def load_bank_csv(self):
       self.bank_csv, temp = load_csv_file()
       self.bank_file_name = os.path.basename(temp)
       if self.bank_csv is not None:
          month_pattern = r'\b(January|February|March|April|May|June|July|August|September|October|November|December)\b'
          match1 = re.search(month_pattern, self.our_file_name, re.IGNORECASE)
          match2 = re.search(month_pattern, self.bank_file_name, re.IGNORECASE)
          if match1 and match2 and match1.group(0) == match2.group(0):
             print("test")
             self.file_month=match1.group(0)
          self.display_columns(self.bank_csv, 'bank')
          self.check_files_loaded()

    def check_files_loaded(self):
        if self.our_csv is not None and self.bank_csv is not None:
            self.match_btn.config(state=tk.NORMAL)

    def match_csvs(self):
        our_columns = {entry.get(): entry.get() for entry, _ in self.our_entries}
        bank_columns = {entry.get(): entry.get() for entry, _ in self.bank_entries}

        matched_df = create_matchbooks(self.our_csv, self.bank_csv, our_columns, bank_columns, True)
        save_to_excel_with_color(matched_df, f"Matched_Output_OurCredits-TheirDebits_{self.file_month}.xlsx")

        matched_df = create_matchbooks(self.our_csv, self.bank_csv, our_columns, bank_columns, False)
        save_to_excel_with_color(matched_df, f"Matched_Output_OurDebits-TheirCredits_{self.file_month}.xlsx")

    def display_columns(self, df, col_type):
        columns = df.columns.tolist()

        # Filter out specific columns
        filtered_columns = [col for col in columns if col.lower() not in ['debits', 'credits', 'debit', 'credit']]

        if col_type == 'our':
            for entry, delete_btn in self.our_entries:
                entry.destroy()
                delete_btn.destroy()
            self.our_entries.clear()

            for idx, col in enumerate(filtered_columns):
                self.add_column_entry(col, idx, 'our')

        else:
            for entry, delete_btn in self.bank_entries:
                entry.destroy()
                delete_btn.destroy()
            self.bank_entries.clear()

            for idx, col in enumerate(filtered_columns):
                self.add_column_entry(col, idx, 'bank')

    def add_column_entry(self, col_name, idx, col_type):
        entry = tk.Entry(self.frame)
        entry.insert(0, col_name)
        entry.grid(row=idx + 1, column=0 if col_type == 'our' else 2, padx=5, pady=5)

        delete_btn = tk.Button(self.frame, text="X", command=lambda: self.delete_entry(entry, delete_btn, col_type))
        delete_btn.grid(row=idx + 1, column=1 if col_type == 'our' else 3, padx=5)

        if col_type == 'our':
            self.our_entries.append((entry, delete_btn))
        else:
            self.bank_entries.append((entry, delete_btn))

    def delete_entry(self, entry, delete_btn, col_type):
        if col_type == 'our':
            self.our_entries.remove((entry, delete_btn))
        else:
            self.bank_entries.remove((entry, delete_btn))

        entry.destroy()
        delete_btn.destroy()

        for idx, (e, btn) in enumerate(self.our_entries if col_type == 'our' else self.bank_entries):
            e.grid(row=idx + 1, column=0 if col_type == 'our' else 2, padx=5, pady=5)
            btn.grid(row=idx + 1, column=1 if col_type == 'our' else 3, padx=5)

root = tk.Tk()
app = CSVMatcherApp(root)
root.mainloop()

